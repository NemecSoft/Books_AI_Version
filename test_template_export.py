import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Fill, Protection
from openpyxl.workbook.workbook import Workbook

class TemplateExportTester:
    def __init__(self, template_path):
        # 使用用户指定的模板文件路径
        self.template_path = template_path
        self.output_path = "d:\\AI\\books\\template_test_output.xlsx"
    
    def _generate_test_data(self):
        # 生成测试数据
        data = [
            ["测试标题1", "1. 事件1\n2. 事件2", "1. 简化事件1\n2. 简化事件2", "1. 详细事件1\n2. 详细事件2"],
            ["测试标题2", "1. 事件3", "1. 简化事件3", "1. 详细事件3\n2. 详细事件4"]
        ]
        return pd.DataFrame(data)
    
    def test_template_export(self):
        """测试使用模板导出Excel"""
        print("开始测试模板导出功能...")
        
        try:
            # 检查模板文件是否存在
            if not os.path.exists(self.template_path):
                print(f"错误：模板文件不存在: {self.template_path}")
                return False
            
            print(f"使用模板文件: {self.template_path}")
            
            # 生成测试数据
            df = self._generate_test_data()
            
            # 加载模板文件，data_only=True表示只读取单元格的值而不是公式
            template_book = load_workbook(self.template_path, data_only=True)
            
            # 确保至少有一个可见的工作表
            visible_sheets = [sheet for sheet in template_book.sheetnames if template_book[sheet].sheet_state == 'visible']
            
            if not visible_sheets:
                # 如果没有可见的工作表，创建一个新的
                template_book = Workbook()
                template_sheet_name = 'Sheet1'
                print("警告：模板文件中没有可见的工作表，创建新的工作表")
            else:
                # 使用第一个可见的工作表
                template_sheet_name = visible_sheets[0]
                print(f"使用模板文件中的可见工作表: {template_sheet_name}")
            
            # 获取模板工作表
            template_worksheet = template_book[template_sheet_name]
            print("模板工作表已加载")
            
            # 创建新的工作簿用于导出
            new_book = Workbook()
            new_worksheet = new_book.active
            new_worksheet.title = 'Sheet1'
            print("创建新的Excel工作簿")
            
            # 记录模板信息
            print(f"模板工作表最大行数: {template_worksheet.max_row}")
            print(f"模板工作表最大列数: {template_worksheet.max_column}")
            print(f"模板工作表列尺寸数量: {len(template_worksheet.column_dimensions)}")
            print(f"模板工作表行尺寸数量: {len(template_worksheet.row_dimensions)}")
            
            # 复制列宽设置
            col_width_count = 0
            for column_letter in template_worksheet.column_dimensions:
                try:
                    if column_letter in template_worksheet.column_dimensions:
                        col_dim = template_worksheet.column_dimensions[column_letter]
                        new_worksheet.column_dimensions[column_letter].width = col_dim.width
                        col_width_count += 1
                        print(f"复制列 {column_letter} 的宽度设置为 {col_dim.width}")
                except Exception as e:
                    print(f"复制列宽时出错: {column_letter}, {str(e)}")
            print(f"成功复制 {col_width_count} 列的宽度设置")
            
            # 复制行高设置
            row_height_count = 0
            for row_idx in template_worksheet.row_dimensions:
                try:
                    if row_idx in template_worksheet.row_dimensions:
                        row_dim = template_worksheet.row_dimensions[row_idx]
                        new_worksheet.row_dimensions[row_idx].height = row_dim.height
                        row_height_count += 1
                        print(f"复制行 {row_idx} 的高度设置为 {row_dim.height}")
                except Exception as e:
                    print(f"复制行高时出错: {row_idx}, {str(e)}")
            print(f"成功复制 {row_height_count} 行的高度设置")
            
            print(f"模板工作表最大行数: {template_worksheet.max_row}")
            print(f"模板工作表最大列数: {template_worksheet.max_column}")
            
            # 写入数据并应用样式
            applied_style_count = 0
            for row_idx, row_data in enumerate(df.itertuples(index=False), start=1):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    # 设置单元格值
                    new_cell = new_worksheet.cell(row=row_idx, column=col_idx)
                    new_cell.value = cell_value
                    print(f"写入单元格 ({row_idx},{col_idx}) 的值: {str(cell_value)[:30]}...")
                    
                    # 直接从模板单元格复制样式
                    try:
                        # 检查模板中是否有对应的单元格
                        if row_idx <= template_worksheet.max_row and col_idx <= template_worksheet.max_column:
                            template_cell = template_worksheet.cell(row=row_idx, column=col_idx)
                            if template_cell.has_style:
                                    # 创建新的样式对象，避免StyleProxy问题
                                    # 字体设置
                                    if template_cell.font:
                                        new_font = Font(
                                            name=template_cell.font.name,
                                            size=template_cell.font.size,
                                            bold=template_cell.font.bold,
                                            italic=template_cell.font.italic,
                                            vertAlign=template_cell.font.vertAlign,
                                            underline=template_cell.font.underline,
                                            strike=template_cell.font.strike,
                                            color=template_cell.font.color
                                        )
                                        new_cell.font = new_font
                                    
                                    # 边框设置
                                    if template_cell.border:
                                        # 使用字典方式创建边框，避免属性名称不匹配问题
                                        border_kwargs = {}
                                        
                                        # 基础边框属性
                                        if hasattr(template_cell.border, 'left'):
                                            border_kwargs['left'] = template_cell.border.left
                                        if hasattr(template_cell.border, 'right'):
                                            border_kwargs['right'] = template_cell.border.right
                                        if hasattr(template_cell.border, 'top'):
                                            border_kwargs['top'] = template_cell.border.top
                                        if hasattr(template_cell.border, 'bottom'):
                                            border_kwargs['bottom'] = template_cell.border.bottom
                                        if hasattr(template_cell.border, 'diagonal'):
                                            border_kwargs['diagonal'] = template_cell.border.diagonal
                                        
                                        # 对角线方向属性（处理不同版本）
                                        if hasattr(template_cell.border, 'diagonal_direction'):
                                            border_kwargs['diagonal_direction'] = template_cell.border.diagonal_direction
                                        elif hasattr(template_cell.border, 'diagonalDirection'):
                                            border_kwargs['diagonalDirection'] = template_cell.border.diagonalDirection
                                        
                                        # 创建边框
                                        new_border = Border(**border_kwargs)
                                        new_cell.border = new_border
                                    
                                    # 填充设置
                                    if template_cell.fill:
                                        # 使用字典方式创建填充，避免属性不存在问题
                                        fill_kwargs = {}
                                        
                                        # 检查属性存在性
                                        if hasattr(template_cell.fill, 'fgColor'):
                                            fill_kwargs['fgColor'] = template_cell.fill.fgColor
                                        if hasattr(template_cell.fill, 'bgColor'):
                                            fill_kwargs['bgColor'] = template_cell.fill.bgColor
                                        if hasattr(template_cell.fill, 'patternType'):
                                            fill_kwargs['patternType'] = template_cell.fill.patternType
                                        
                                        # 尝试创建填充
                                        if fill_kwargs:
                                            try:
                                                new_fill = Fill(**fill_kwargs)
                                                new_cell.fill = new_fill
                                            except Exception as e:
                                                print(f"填充样式创建失败: {e}")
                                    
                                    # 其他属性
                                    new_cell.number_format = template_cell.number_format
                                    
                                    if template_cell.protection:
                                        new_protection = Protection(
                                            locked=template_cell.protection.locked,
                                            hidden=template_cell.protection.hidden
                                        )
                                        new_cell.protection = new_protection
                                    
                                    if template_cell.alignment:
                                        new_alignment = Alignment(
                                            horizontal=template_cell.alignment.horizontal,
                                            vertical=template_cell.alignment.vertical,
                                            textRotation=template_cell.alignment.textRotation,
                                            wrapText=template_cell.alignment.wrapText,
                                            shrinkToFit=template_cell.alignment.shrinkToFit,
                                            indent=template_cell.alignment.indent,
                                            relativeIndent=template_cell.alignment.relativeIndent,
                                            justifyLastLine=template_cell.alignment.justifyLastLine,
                                            readingOrder=template_cell.alignment.readingOrder
                                        )
                                        new_cell.alignment = new_alignment
                                    
                                    applied_style_count += 1
                                    print(f"成功应用单元格 ({row_idx},{col_idx}) 的样式")
                    except Exception as e:
                        print(f"应用单元格样式时出错: ({row_idx},{col_idx}), {str(e)}")
                    
                    # 如果没有模板样式或样式复制失败，确保设置了文本换行
                    if hasattr(new_cell, 'alignment') and (not new_cell.alignment or not new_cell.alignment.wrap_text):
                        new_cell.alignment = Alignment(wrap_text=True)
                        print(f"为单元格 ({row_idx},{col_idx}) 设置默认文本换行")
            print(f"成功为 {applied_style_count} 个单元格应用了样式")
            style_cell_count = applied_style_count  # 保持变量一致性
            
            # 保存新工作簿
            new_book.save(self.output_path)
            print(f"成功保存工作簿到: {self.output_path}")
            
            # 检查文件是否存在并验证大小
            if os.path.exists(self.output_path):
                file_size = os.path.getsize(self.output_path)
                print(f"文件大小: {file_size} 字节")
                
                # 重新读取文件验证数据
                df_read = pd.read_excel(self.output_path, header=None)
                print(f"导出文件包含 {len(df_read)} 行和 {len(df_read.columns)} 列")
                
                # 验证第一个单元格的值
                if df_read.iloc[0, 0] == df.iloc[0, 0]:
                    print("数据验证成功：第一个单元格值正确")
                    print("\n测试结果：成功 ✓")
                    print(f"您可以查看导出的测试文件: {self.output_path}")
                    return True
                else:
                    print("数据验证失败：第一个单元格值不正确")
                    print("\n测试结果：失败 ✗")
                    return False
            else:
                print("错误：导出文件不存在")
                print("\n测试结果：失败 ✗")
                return False
                
        except Exception as e:
            import traceback
            print(f"测试过程中出错: {str(e)}")
            print("错误详情:")
            traceback.print_exc()
            print("\n测试结果：失败 ✗")
            return False

if __name__ == "__main__":
    # 使用用户提供的模板文件路径
    template_path = "d:\\AI\\books\\白眉大侠\\章回\\事件列表\\test.xlsx"
    
    tester = TemplateExportTester(template_path)
    tester.test_template_export()
