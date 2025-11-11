import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import json
import os
import pandas as pd
from datetime import datetime
import types  # 在文件顶部导入types模块
from openpyxl import load_workbook, Workbook  # 导入openpyxl用于处理Excel模板
# 完整导入所有需要的样式类
from openpyxl.styles import (
    Alignment, Font, Border, Side,
    Fill, Protection, PatternFill, 
    GradientFill, Color, colors
)  # 导入所有需要的样式类

class JSONConverterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("JSON转Excel/Markdown工具")
        self.root.geometry("800x600")
        self.root.resizable(True, True)
        
        # 设置中文字体支持
        try:
            self.style = ttk.Style()
            self.style.configure("TButton", font=('SimHei', 10))
            self.style.configure("TLabel", font=('SimHei', 10))
            print("成功设置中文字体")
        except Exception as e:
            print(f"设置字体时出错: {e}")
        
        # 数据存储
        self.json_data = []
        self.file_paths = []
        
        # 创建界面
        self.create_widgets()

    def create_widgets(self):
        # 创建主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 按钮区域
        button_frame = ttk.Frame(main_frame, padding="5")
        button_frame.pack(fill=tk.X, side=tk.TOP)
        
        self.add_file_btn = ttk.Button(button_frame, text="添加JSON文件", command=self.add_json_files)
        self.add_file_btn.pack(side=tk.LEFT, padx=5)
        
        self.add_folder_btn = ttk.Button(button_frame, text="添加文件夹中的JSON", command=self.add_folder_json)
        self.add_folder_btn.pack(side=tk.LEFT, padx=5)
        
        self.remove_btn = ttk.Button(button_frame, text="移除选中", command=self.remove_selected)
        self.remove_btn.pack(side=tk.LEFT, padx=5)
        
        self.clear_btn = ttk.Button(button_frame, text="清空列表", command=self.clear_all)
        self.clear_btn.pack(side=tk.LEFT, padx=5)
        
        self.sort_label = ttk.Label(button_frame, text="排序方式:")
        self.sort_label.pack(side=tk.LEFT, padx=5, pady=5)
        
        self.sort_var = tk.StringVar(value="文件名")
        self.sort_combo = ttk.Combobox(button_frame, textvariable=self.sort_var, values=["文件名", "标题"], state="readonly", width=10)
        self.sort_combo.pack(side=tk.LEFT, padx=5)
        
        self.sort_btn = ttk.Button(button_frame, text="执行排序", command=self.sort_data)
        self.sort_btn.pack(side=tk.LEFT, padx=5)
        
        # 导出按钮
        export_frame = ttk.Frame(main_frame, padding="5")
        export_frame.pack(fill=tk.X, side=tk.TOP)
        
        self.to_excel_btn = ttk.Button(export_frame, text="导出为Excel", command=self.export_to_excel)
        self.to_excel_btn.pack(side=tk.LEFT, padx=5)
        
        self.to_csv_btn = ttk.Button(export_frame, text="导出为CSV", command=self.export_to_csv)
        self.to_csv_btn.pack(side=tk.LEFT, padx=5)
        
        self.to_markdown_btn = ttk.Button(export_frame, text="导出为Markdown", command=self.export_to_markdown)
        self.to_markdown_btn.pack(side=tk.LEFT, padx=5)
        
        # 状态标签
        self.status_var = tk.StringVar(value="就绪")
        status_label = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_label.pack(side=tk.BOTTOM, fill=tk.X)
        
        # 创建表格视图
        columns = ("文件路径", "标题", "状态")
        self.tree = ttk.Treeview(main_frame, columns=columns, show="headings")
        
        # 设置列宽
        self.tree.column("文件路径", width=400)
        self.tree.column("标题", width=250)
        self.tree.column("状态", width=100)
        
        # 设置列标题
        for col in columns:
            self.tree.heading(col, text=col)
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        
        # 放置表格和滚动条
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.pack(fill=tk.BOTH, expand=True, side=tk.TOP)

    def add_json_files(self):
        """添加JSON文件"""
        file_types = [("JSON文件", "*.json"), ("所有文件", "*.*")]
        file_paths = filedialog.askopenfilenames(title="选择JSON文件", filetypes=file_types)
        
        if file_paths:
            self.process_json_files(file_paths)

    def add_folder_json(self):
        """添加文件夹中的所有JSON文件"""
        folder_path = filedialog.askdirectory(title="选择包含JSON文件的文件夹")
        
        if folder_path:
            file_paths = []
            for root, _, files in os.walk(folder_path):
                for file in files:
                    if file.lower().endswith('.json'):
                        file_paths.append(os.path.join(root, file))
            
            if file_paths:
                self.process_json_files(file_paths)
            else:
                messagebox.showinfo("提示", f"文件夹 '{folder_path}' 中未找到JSON文件")

    def process_json_files(self, file_paths):
        """处理JSON文件并添加到列表中"""
        success_count = 0
        error_count = 0
        
        for file_path in file_paths:
            if file_path in self.file_paths:
                continue  # 跳过已添加的文件
            
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                # 验证JSON结构
                if "基础信息" not in data or "标题" not in data["基础信息"]:
                    raise ValueError("JSON结构不符合要求，缺少必要字段")
                
                # 提取必要信息
                title = data["基础信息"]["标题"]
                
                # 检查其他必要字段
                if "极简版" not in data or "简化版" not in data or "详细版" not in data:
                    raise ValueError("JSON结构不符合要求，缺少必要的事件列表字段")
                
                # 防御性处理：确保所有列表字段都不是生成器
                极简版_data = data["极简版"]
                if isinstance(极简版_data, types.GeneratorType) or (hasattr(极简版_data, '__iter__') and not hasattr(极简版_data, '__len__') and not isinstance(极简版_data, (str, bytes, dict))):
                    极简版_data = list(极简版_data)
                elif not isinstance(极简版_data, list):
                    极简版_data = [极简版_data] if 极简版_data is not None else []
                    
                简化版_data = data["简化版"]
                if isinstance(简化版_data, types.GeneratorType) or (hasattr(简化版_data, '__iter__') and not hasattr(简化版_data, '__len__') and not isinstance(简化版_data, (str, bytes, dict))):
                    简化版_data = list(简化版_data)
                elif not isinstance(简化版_data, list):
                    简化版_data = [简化版_data] if 简化版_data is not None else []
                    
                详细版_data = data["详细版"]
                if isinstance(详细版_data, types.GeneratorType) or (hasattr(详细版_data, '__iter__') and not hasattr(详细版_data, '__len__') and not isinstance(详细版_data, (str, bytes, dict))):
                    详细版_data = list(详细版_data)
                elif not isinstance(详细版_data, list):
                    详细版_data = [详细版_data] if 详细版_data is not None else []
                    
                # 添加到数据列表
                self.json_data.append({
                    "file_path": file_path,
                    "title": title,
                    "极简版": 极简版_data,
                    "简化版": 简化版_data,
                    "详细版": 详细版_data
                })
                
                self.file_paths.append(file_path)
                self.tree.insert("", tk.END, values=(os.path.basename(file_path), title, "成功"))
                success_count += 1
                
            except json.JSONDecodeError:
                self.tree.insert("", tk.END, values=(os.path.basename(file_path), "", "JSON格式错误"))
                error_count += 1
            except Exception as e:
                print(f"处理文件 {file_path} 时出错: {str(e)}")
                self.tree.insert("", tk.END, values=(os.path.basename(file_path), "", f"错误: {str(e)}"))
                error_count += 1
        
        self.status_var.set(f"添加完成 - 成功: {success_count}, 失败: {error_count}")

    def remove_selected(self):
        """移除选中的项目"""
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showinfo("提示", "请先选择要移除的项目")
            return
        
        removed_count = 0
        for item in selected_items:
            file_name = self.tree.item(item, "values")[0]
            # 查找并移除对应的文件路径和数据
            for i, data in enumerate(self.json_data):
                if os.path.basename(data["file_path"]) == file_name:
                    del self.json_data[i]
                    self.file_paths.remove(data["file_path"])
                    removed_count += 1
                    break
            # 从树视图中删除
            self.tree.delete(item)
        
        self.status_var.set(f"已移除 {removed_count} 个项目")

    def clear_all(self):
        """清空所有数据"""
        if messagebox.askyesno("确认", "确定要清空所有数据吗？"):
            self.tree.delete(*self.tree.get_children())
            self.json_data = []
            self.file_paths = []
            self.status_var.set("已清空所有数据")

    def sort_data(self):
        """根据选择的方式排序数据"""
        if not self.json_data:
            messagebox.showinfo("提示", "没有数据可排序")
            return
        
        sort_by = self.sort_var.get()
        if sort_by == "文件名":
            self.json_data.sort(key=lambda x: os.path.basename(x["file_path"]))
        elif sort_by == "标题":
            self.json_data.sort(key=lambda x: x["标题"])
        
        # 更新树视图
        self.tree.delete(*self.tree.get_children())
        for data in self.json_data:
            self.tree.insert("", tk.END, values=(os.path.basename(data["file_path"]), data["标题"], "成功"))
        
        self.status_var.set(f"已按 {sort_by} 排序")

    def export_to_excel(self):
        """导出为Excel文件，支持使用模板文件格式"""
        try:
            # 防御性编程：确保json_data是列表类型
            if isinstance(self.json_data, types.GeneratorType):
                self.json_data = list(self.json_data)
            elif hasattr(self.json_data, '__iter__') and not hasattr(self.json_data, '__len__') and not isinstance(self.json_data, (str, bytes, dict)):
                self.json_data = list(self.json_data)
            if not isinstance(self.json_data, list):
                self.json_data = [self.json_data] if self.json_data is not None else []
            
            if not self.json_data:
                messagebox.showinfo("提示", "没有数据可导出")
                return
            
            # 询问用户是否使用模板文件
            use_template = messagebox.askyesno("使用模板", "是否使用现有Excel文件作为模板？")
            template_path = None
            
            if use_template:
                template_path = filedialog.askopenfilename(
                    title="选择模板Excel文件",
                    filetypes=[("Excel文件", "*.xlsx")]
                )
                if not template_path:
                    # 用户取消选择模板，使用默认格式
                    use_template = False
            
            # 创建数据框
            data_to_export = []
            
            for data in self.json_data:
                if not isinstance(data, dict):
                    continue
                
                # 处理各字段
                极简版_list = self._safe_convert_to_list(data.get("极简版", []))
                简化版_list = self._safe_convert_to_list(data.get("简化版", []))
                详细版_list = self._safe_convert_to_list(data.get("详细版", []))
                
                # 生成字符串内容
                极简版_str = "\n".join([f"{i}. {str(event)}" for i, event in enumerate(极简版_list, 1)] if 极简版_list else [])
                简化版_str = "\n".join([f"{i}. {str(event)}" for i, event in enumerate(简化版_list, 1)] if 简化版_list else [])
                详细版_str = "\n".join([f"{i}. {str(event)}" for i, event in enumerate(详细版_list, 1)] if 详细版_list else [])
                
                # 确保标题不为空，并添加详细调试信息
                # 从self.json_data中，标题字段应该是"title"而不是"标题"
                标题内容 = data.get("title", "")
                print(f"添加数据: 标题='{标题内容}', title存在={"title" in data}, title类型={type(标题内容)}")
                print(f"数据结构: {list(data.keys())}")
                
                data_to_export.append([
                    标题内容,
                    极简版_str,
                    简化版_str,
                    详细版_str
                ])
            
            # 创建DataFrame，不设置列名
            df = pd.DataFrame(data_to_export)            
            # 打印前几行数据以调试
            print(f"DataFrame形状: {df.shape}")
            print("DataFrame前几行:")
            print(df.head())
            
            default_filename = f"小说事件汇总_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx")],
                initialfile=default_filename
            )
            
            if file_path:
                if use_template and template_path:
                    try:
                        # 使用模板文件的格式
                        from openpyxl import load_workbook
                        from openpyxl.styles import Alignment
                        
                        # 加载模板文件，data_only=True表示只读取单元格的值而不是公式
                        template_book = load_workbook(template_path, data_only=True)
                        
                        # 确保至少有一个可见的工作表
                        visible_sheets = [sheet for sheet in template_book.sheetnames if template_book[sheet].sheet_state == 'visible']
                        
                        if not visible_sheets:
                            # 如果没有可见的工作表，创建一个新的
                            from openpyxl.workbook.workbook import Workbook
                            template_book = Workbook()
                            template_sheet_name = 'Sheet1'
                        else:
                            # 使用第一个可见的工作表
                            template_sheet_name = visible_sheets[0]
                        
                        # 获取模板工作表
                        template_worksheet = template_book[template_sheet_name]
                        
                        # 直接使用模板工作簿，不创建新工作簿
                        new_book = template_book
                        new_worksheet = template_book[template_sheet_name]
                        
                        # 导入所有必要的样式类
                        from openpyxl.styles import Font, Border, PatternFill, Protection, Alignment, Side
                        
                        # 在清除内容前，先保存第一行的样式
                        print("保存第一行样式...")
                        # 存储第一行的样式对象，而不是单元格引用
                        first_row_styles = []
                        max_col = template_worksheet.max_column
                        # 确保至少保存4列的样式（我们需要的列数）
                        for col_idx in range(1, max(5, max_col + 1)):  # 保存至少4列样式
                            try:
                                # 获取第一行单元格
                                template_cell = template_worksheet.cell(row=1, column=col_idx)
                                # 保存样式的副本，而不是引用
                                style_copy = {}
                                
                                # 单独处理每个样式属性，确保正确复制
                                if template_cell.font:
                                    style_copy['font'] = Font(
                                        name=template_cell.font.name if hasattr(template_cell.font, 'name') else 'Arial',
                                        size=template_cell.font.size if hasattr(template_cell.font, 'size') else 11,
                                        bold=template_cell.font.bold if hasattr(template_cell.font, 'bold') else False,
                                        italic=template_cell.font.italic if hasattr(template_cell.font, 'italic') else False,
                                        color=template_cell.font.color if hasattr(template_cell.font, 'color') else None,
                                        underline=template_cell.font.underline if hasattr(template_cell.font, 'underline') else 'none'
                                    )
                                else:
                                    style_copy['font'] = Font()
                                
                                # 处理边框
                                if template_cell.border:
                                    # 创建新的Side对象而不是调用copy()
                                    def create_side(side_obj):
                                        if side_obj and hasattr(side_obj, 'style') and hasattr(side_obj, 'color'):
                                            return Side(
                                                style=side_obj.style,
                                                color=side_obj.color
                                            )
                                        return Side()
                                    
                                    style_copy['border'] = Border(
                                        left=create_side(template_cell.border.left) if hasattr(template_cell.border, 'left') else Side(),
                                        right=create_side(template_cell.border.right) if hasattr(template_cell.border, 'right') else Side(),
                                        top=create_side(template_cell.border.top) if hasattr(template_cell.border, 'top') else Side(),
                                        bottom=create_side(template_cell.border.bottom) if hasattr(template_cell.border, 'bottom') else Side(),
                                        diagonal=template_cell.border.diagonal if hasattr(template_cell.border, 'diagonal') else None,
                                        diagonal_direction=template_cell.border.diagonal_direction if hasattr(template_cell.border, 'diagonal_direction') else None,
                                        outline=template_cell.border.outline if hasattr(template_cell.border, 'outline') else False,
                                        vertical=template_cell.border.vertical if hasattr(template_cell.border, 'vertical') else None,
                                        horizontal=template_cell.border.horizontal if hasattr(template_cell.border, 'horizontal') else None
                                    )
                                else:
                                    style_copy['border'] = Border()
                                
                                # 处理填充
                                if template_cell.fill:
                                    style_copy['fill'] = PatternFill(
                                        patternType=template_cell.fill.patternType if hasattr(template_cell.fill, 'patternType') else None,
                                        fgColor=template_cell.fill.fgColor if hasattr(template_cell.fill, 'fgColor') else None,
                                        bgColor=template_cell.fill.bgColor if hasattr(template_cell.fill, 'bgColor') else None
                                    )
                                else:
                                    style_copy['fill'] = PatternFill()
                                
                                # 复制其他样式属性
                                style_copy['number_format'] = template_cell.number_format
                                
                                if template_cell.protection:
                                    style_copy['protection'] = Protection(
                                        locked=template_cell.protection.locked if hasattr(template_cell.protection, 'locked') else True,
                                        hidden=template_cell.protection.hidden if hasattr(template_cell.protection, 'hidden') else False
                                    )
                                else:
                                    style_copy['protection'] = Protection()
                                
                                if template_cell.alignment:
                                    style_copy['alignment'] = Alignment(
                                        horizontal=template_cell.alignment.horizontal if hasattr(template_cell.alignment, 'horizontal') else 'general',
                                        vertical=template_cell.alignment.vertical if hasattr(template_cell.alignment, 'vertical') else 'bottom',
                                        textRotation=template_cell.alignment.textRotation if hasattr(template_cell.alignment, 'textRotation') else 0,
                                        wrapText=True,  # 强制设置文本换行
                                        shrinkToFit=template_cell.alignment.shrinkToFit if hasattr(template_cell.alignment, 'shrinkToFit') else False,
                                        indent=template_cell.alignment.indent if hasattr(template_cell.alignment, 'indent') else 0
                                    )
                                else:
                                    style_copy['alignment'] = Alignment(wrapText=True)
                                
                                first_row_styles.append(style_copy)
                                print(f"  保存第{col_idx}列样式: {style_copy['font'].name}")
                            except Exception as e:
                                print(f"  保存第{col_idx}列样式时出错: {str(e)}")
                                # 添加默认样式作为后备
                                first_row_styles.append({
                                    'font': Font(),
                                    'border': Border(),
                                    'fill': PatternFill(),
                                    'number_format': 'General',
                                    'protection': Protection(),
                                    'alignment': Alignment(wrapText=True)
                                })
                        
                        # 保存第一行样式后，我们不清除整个工作表，而是直接在写入时覆盖单元格值
                        # 这样可以避免清除操作对样式的影响
                        print("保留模板工作表的原始样式结构，准备写入新数据...")
                        
                        # 直接写入数据，不写标题行
                        print(f"使用模板格式导出，数据形状: {df.shape}")
                        
                        # 预先处理数据，确保每行都有完整的四列数据
                        processed_data = []
                        for row_data in df.itertuples(index=False):
                            row_list = list(row_data)
                            # 确保每行都有四列数据，不足的用空字符串填充
                            while len(row_list) < 4:
                                row_list.append("")
                            # 截断过长的行
                            if len(row_list) > 4:
                                row_list = row_list[:4]
                            processed_data.append(row_list)
                            print(f"处理后的数据行: {row_list}")
                            print(f"  第一列值类型: {type(row_list[0])}, 值: {row_list[0] if row_list[0] else '空字符串'}")
                        
                        # 强制从第1行开始写入数据
                        print("======= 关键调试信息 =======")
                        print("这里是行号设置的关键位置")
                        start_row = 1
                        print(f"明确设置start_row = {start_row}")
                        print(f"工作表最大行数: {new_worksheet.max_row}")
                        print(f"工作表最大列数: {new_worksheet.max_column}")
                        print("======= 关键调试信息结束 =======")
                        
                        # 修复第一列数据导出问题
                        print(f"准备写入数据，processed_data长度: {len(processed_data)}")
                        print(f"当前start_row值: {start_row}")
                        print("======= 开始写入数据循环 =======")
                        
                        # 特殊处理：确保第1行被写入
                        if processed_data:
                            print("======= 特殊处理：直接写入第1行 =======")
                            row_list = processed_data[0]
                            row_idx = 1
                            print(f"直接写入行号{row_idx}，数据类型: {type(row_list[0])}")
                            print(f"第一行数据内容: {str(row_list[0])[:20]}...")
                            
                            # 确保所有四列都被写入，特别是第一列（标题列）
                            for col_idx in range(1, 5):  # 强制写入4列
                                # 特别处理第一列，确保使用正确的标题数据
                                if col_idx == 1:
                                    # 直接使用row_list中的第一列数据（来自data["title"]）
                                    if (col_idx - 1) < len(row_list) and row_list[col_idx - 1]:
                                        cell_value = row_list[col_idx - 1]
                                        print(f"  第一行 - 使用正确的标题数据: {cell_value[:20]}...")
                                    else:
                                        cell_value = "事件标题"
                                        print(f"  第一行 - 设置默认标题: {cell_value}")
                                else:
                                    cell_value = row_list[col_idx - 1] if (col_idx - 1) < len(row_list) else ""
                                
                                # 获取或创建单元格
                                new_cell = new_worksheet.cell(row=row_idx, column=col_idx)
                                
                                # 设置单元格值
                                new_cell.value = cell_value
                                print(f"  第一行 - 写入单元格({row_idx},{col_idx}): {cell_value[:20] if isinstance(cell_value, str) else cell_value}")
                                
                                # 确保第一行始终有样式
                                try:
                                    # 优先尝试使用保存的样式
                                    if col_idx <= len(first_row_styles):
                                        # 获取保存的样式
                                        saved_style = first_row_styles[col_idx - 1]
                                        
                                        # 应用所有样式属性
                                        # 创建新的Font对象以确保样式正确应用
                                        if hasattr(saved_style['font'], 'name'):
                                            font_obj = Font(
                                                name=saved_style['font'].name,
                                                size=saved_style['font'].size if hasattr(saved_style['font'], 'size') else 11,
                                                bold=saved_style['font'].bold if hasattr(saved_style['font'], 'bold') else False,
                                                italic=saved_style['font'].italic if hasattr(saved_style['font'], 'italic') else False,
                                                color=saved_style['font'].color if hasattr(saved_style['font'], 'color') else None
                                            )
                                            new_cell.font = font_obj
                                            print(f"  第一行 - 应用字体样式到单元格({row_idx},{col_idx})")
                                        
                                        # 创建新的Border对象
                                        if hasattr(saved_style['border'], 'left'):
                                            border_obj = Border(
                                                left=saved_style['border'].left,
                                                right=saved_style['border'].right,
                                                top=saved_style['border'].top,
                                                bottom=saved_style['border'].bottom
                                            )
                                            new_cell.border = border_obj
                                            print(f"  第一行 - 应用边框样式到单元格({row_idx},{col_idx})")
                                        
                                        # 创建新的Fill对象
                                        if hasattr(saved_style['fill'], 'patternType'):
                                            fill_obj = PatternFill(
                                                start_color=saved_style['fill'].start_color,
                                                end_color=saved_style['fill'].end_color,
                                                patternType=saved_style['fill'].patternType
                                            )
                                            new_cell.fill = fill_obj
                                            print(f"  第一行 - 应用填充样式到单元格({row_idx},{col_idx})")
                                        
                                        # 应用对齐方式 - 确保文本换行
                                        alignment_obj = Alignment(
                                            horizontal=saved_style['alignment'].horizontal if hasattr(saved_style['alignment'], 'horizontal') else None,
                                            vertical=saved_style['alignment'].vertical if hasattr(saved_style['alignment'], 'vertical') else None,
                                            wrapText=True  # 确保文本换行
                                        )
                                        new_cell.alignment = alignment_obj
                                        
                                        # 应用其他样式属性
                                        new_cell.number_format = saved_style['number_format']
                                        new_cell.protection = saved_style['protection']
                                        
                                        print(f"  第一行 - 完成应用第{col_idx}列样式到单元格({row_idx},{col_idx})")
                                    else:
                                        # 没有保存的样式，直接设置默认样式
                                        raise Exception("没有保存的样式，使用默认样式")
                                except Exception as e:
                                    print(f"  第一行 - 应用样式到单元格({row_idx},{col_idx})时出错或没有保存的样式: {str(e)}")
                                    
                                    # 手动设置默认样式作为备选方案
                                    try:
                                        # 手动设置默认样式
                                        new_cell.font = Font(
                                            name='SimHei',  # 使用中文字体
                                            size=12,
                                            bold=True
                                        )
                                        # 确保文本换行
                                        new_cell.alignment = Alignment(
                                            horizontal='center',
                                            vertical='center',
                                            wrapText=True
                                        )
                                        # 设置边框
                                        side = Side(style='thin', color=colors.BLACK)
                                        new_cell.border = Border(left=side, right=side, top=side, bottom=side)
                                        # 设置填充色为浅灰色
                                        new_cell.fill = PatternFill(
                                            start_color='F2F2F2', 
                                            end_color='F2F2F2', 
                                            patternType='solid'
                                        )
                                        print(f"  第一行 - 成功应用默认样式到单元格({row_idx},{col_idx})")
                                    except Exception as alt_e:
                                        print(f"  第一行 - 应用默认样式也失败: {str(alt_e)}")
                        
                        # 然后写入剩余的数据行（从第2行开始）
                        print("======= 写入剩余数据行 =======")
                        for idx in range(1, len(processed_data)):
                            row_idx = idx + 1  # 行号从2开始
                            row_list = processed_data[idx]
                            print(f"写入索引{idx}，行号{row_idx}，数据预览: {str(row_list[0])[:20]}...")
                               
                            # 确保所有四列都被写入，特别是第一列（标题列）
                            for col_idx in range(1, 5):  # 强制写入4列
                                # 特别处理第一列，确保使用正确的标题数据
                                if col_idx == 1:
                                    # 直接使用row_list中的第一列数据（来自data["title"]）
                                    if (col_idx - 1) < len(row_list) and row_list[col_idx - 1]:
                                        cell_value = row_list[col_idx - 1]
                                        print(f"  使用正确的标题数据: {cell_value[:20]}...")
                                    else:
                                        cell_value = f"事件{row_idx}"
                                        print(f"  警告: 第{row_idx}行第一列数据为空，设置默认值: {cell_value}")
                                else:
                                    cell_value = row_list[col_idx - 1] if (col_idx - 1) < len(row_list) else ""
                                
                                # 获取或创建单元格
                                new_cell = new_worksheet.cell(row=row_idx, column=col_idx)
                                
                                # 设置单元格值
                                new_cell.value = cell_value
                                print(f"  写入单元格({row_idx},{col_idx}): {cell_value[:20] if isinstance(cell_value, str) else cell_value}")
                                
                                # 从保存的样式中应用到所有行
                                if col_idx <= len(first_row_styles):
                                    # 获取保存的样式
                                    saved_style = first_row_styles[col_idx - 1]
                                    
                                    # 应用所有样式属性
                                    try:
                                        # 直接应用字体样式
                                        new_cell.font = saved_style['font']
                                        print(f"  应用字体样式到单元格({row_idx},{col_idx})")
                                        
                                        # 直接应用边框样式
                                        new_cell.border = saved_style['border']
                                        print(f"  应用边框样式到单元格({row_idx},{col_idx})")
                                        
                                        # 直接应用填充样式
                                        new_cell.fill = saved_style['fill']
                                        print(f"  应用填充样式到单元格({row_idx},{col_idx})")
                                        
                                        # 应用其他样式属性
                                        new_cell.number_format = saved_style['number_format']
                                        new_cell.protection = saved_style['protection']
                                        new_cell.alignment = saved_style['alignment']
                                        
                                        print(f"  完成应用第{col_idx}列样式到单元格({row_idx},{col_idx})")
                                    except Exception as e:
                                        print(f"  应用样式到单元格({row_idx},{col_idx})时出错: {str(e)}")
                                        
                                        # 尝试手动设置关键样式属性作为备选方案
                                        try:
                                            # 手动设置字体
                                            if hasattr(saved_style['font'], 'name'):
                                                new_cell.font = Font(
                                                    name=saved_style['font'].name,
                                                    size=saved_style['font'].size if hasattr(saved_style['font'], 'size') else 11,
                                                    bold=saved_style['font'].bold if hasattr(saved_style['font'], 'bold') else False
                                                )
                                            # 确保文本换行
                                            new_cell.alignment = Alignment(wrapText=True)
                                            print(f"  应用备选样式成功")
                                        except Exception:
                                            pass
                                else:
                                    # 如果没有保存的样式，至少确保设置文本换行
                                    try:
                                        new_cell.alignment = Alignment(wrapText=True)
                                        print(f"  为单元格({row_idx},{col_idx})设置默认文本换行")
                                    except Exception as e:
                                        print(f"  设置文本换行时出错: ({row_idx},{col_idx}), {str(e)}")
                        
                        # 保存新工作簿
                        new_book.save(file_path)
                        
                        print("成功使用模板导出Excel文件")
                    except Exception as e:
                        print(f"使用模板时出错: {str(e)}")
                        messagebox.showwarning("警告", f"使用模板时出错: {str(e)}，将使用默认格式")
                        # 回退到默认格式
                        use_template = False
                else:
                    # 使用默认格式
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        # 为数据框设置明确的列名，并确保即使header=False也能正确写入所有列
                        df.columns = ['标题', '极简版', '简化版', '详细版']
                        print(f"数据框内容前5行:\n{df.head()}")
                        print(f"数据框列信息: {df.info()}")
                        
                        # 创建一个新的数据框，显式确保所有四列都有值，特别处理第一列
                        fixed_data = []
                        for row_idx, row in enumerate(df.iterrows(), start=1):
                            index, row_values = row
                            # 确保每一行都有四列数据，缺失的列填充空字符串
                            fixed_row = []
                            for i in range(4):  # 处理四列
                                if i == 0:  # 第一列特殊处理
                                    if len(row_values) > 0 and row_values.iloc[0]:
                                        fixed_row.append(row_values.iloc[0])
                                        print(f"  使用正确的标题数据: {row_values.iloc[0][:20]}...")
                                    else:
                                        # 直接使用默认值
                                        fixed_row.append(f"事件{row_idx}")
                                        print(f"  警告: 第{row_idx}行第一列数据为空，设置默认值: 事件{row_idx}")
                                elif i < len(row_values):
                                    fixed_row.append(row_values.iloc[i])
                                else:
                                    fixed_row.append("")
                            print(f"默认导出处理后的数据行: {fixed_row}")
                            fixed_data.append(fixed_row)
                        
                        # 创建新的数据框并写入Excel
                        fixed_df = pd.DataFrame(fixed_data, columns=['标题', '极简版', '简化版', '详细版'])
                        print(f"修复后的数据框形状: {fixed_df.shape}")
                        print(f"修复后的数据框前5行:\n{fixed_df.head()}")
                        
                        # 直接使用to_excel，确保所有列都被写入
                        fixed_df.to_excel(writer, index=False, header=False)
                        print(f"使用默认格式导出，数据形状: {df.shape}, 列名: {df.columns.tolist()}")
                        
                        # 优化格式
                        worksheet = writer.sheets['Sheet1']
                        from openpyxl.styles import Font
                        data_font = Font(size=12)
                        header_font = Font(size=12, bold=True)
                        
                        # 转换为列表
                        rows_list = list(worksheet.rows)
                        max_lines_per_row = [0] * (len(rows_list) + 1)
                        max_widths = [0] * len(rows_list[0]) if rows_list else []
                        
                        # 设置字体并计算宽高
                        for row_idx, row in enumerate(worksheet.iter_rows(), start=1):
                            for i, cell in enumerate(row):
                                if cell.value is not None:
                                    # 为标题行设置粗体字体，数据行设置普通字体
                                    if row_idx == 1:
                                        cell.font = header_font
                                    else:
                                        cell.font = data_font
                                    try:
                                        lines = str(cell.value).split('\n')
                                        num_lines = len(lines)
                                        max_lines_per_row[cell.row] = max(max_lines_per_row[cell.row], num_lines)
                                        
                                        # 计算列宽
                                        max_line_length = max([len(line) for line in lines] + [0])
                                        max_widths[i] = max(max_widths[i], max_line_length)
                                    except:
                                        max_lines_per_row[cell.row] = max(max_lines_per_row[cell.row], 1)
                                        max_widths[i] = max(max_widths[i], 10)
                        
                        # 设置列宽
                        for i, width in enumerate(max_widths):
                            column_letter = chr(65 + i)
                            worksheet.column_dimensions[column_letter].width = min(width * 1.1 + 2, 200)
                        
                        # 设置行高
                        for row_idx in range(1, len(max_lines_per_row)):
                            if max_lines_per_row[row_idx] > 0:
                                worksheet.row_dimensions[row_idx].height = 15 * max_lines_per_row[row_idx]
                
                self.status_var.set(f"成功导出到: {file_path}")
                messagebox.showinfo("成功", f"已成功导出到: {file_path}")
        except Exception as e:
            print(f"导出Excel失败: {str(e)}")
            messagebox.showerror("错误", f"导出失败: {str(e)}")
            self.status_var.set(f"导出失败: {str(e)}")
    
    def _safe_convert_to_list(self, data):
        """安全地将数据转换为列表"""
        try:
            if isinstance(data, types.GeneratorType):
                return list(data)
            elif hasattr(data, '__iter__') and not isinstance(data, (str, bytes, dict)):
                return list(data)
            else:
                return [data] if data is not None else []
        except:
            return []

    def export_to_csv(self):
        """导出为CSV文件"""
        try:
            # 防御性编程：确保json_data是列表类型
            if isinstance(self.json_data, types.GeneratorType):
                self.json_data = list(self.json_data)
            elif hasattr(self.json_data, '__iter__') and not hasattr(self.json_data, '__len__') and not isinstance(self.json_data, (str, bytes, dict)):
                self.json_data = list(self.json_data)
            if not isinstance(self.json_data, list):
                self.json_data = [self.json_data] if self.json_data is not None else []
            
            if not self.json_data:
                messagebox.showinfo("提示", "没有数据可导出")
                return
            
            # 创建数据框
            data_to_export = []
            
            for data in self.json_data:
                if not isinstance(data, dict):
                    continue
                
                # 处理各字段
                极简版_list = self._safe_convert_to_list(data.get("极简版", []))
                简化版_list = self._safe_convert_to_list(data.get("简化版", []))
                详细版_list = self._safe_convert_to_list(data.get("详细版", []))
                
                # 生成字符串内容（CSV中用;分隔多行内容）
                极简版_str = "; ".join([f"{i}. {str(event)}" for i, event in enumerate(极简版_list, 1)] if 极简版_list else [])
                简化版_str = "; ".join([f"{i}. {str(event)}" for i, event in enumerate(简化版_list, 1)] if 简化版_list else [])
                详细版_str = "; ".join([f"{i}. {str(event)}" for i, event in enumerate(详细版_list, 1)] if 详细版_list else [])
                
                data_to_export.append([
                    data.get("标题", ""),
                    极简版_str,
                    简化版_str,
                    详细版_str
                ])
            
            # 创建DataFrame并导出
            df = pd.DataFrame(data_to_export, columns=["书籍名称", "极简版", "简化版", "详细版"])
            
            default_filename = f"小说事件汇总_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            file_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV文件", "*.csv")],
                initialfile=default_filename
            )
            
            if file_path:
                # 使用utf-8-sig编码以支持Excel正确识别中文
                df.to_csv(file_path, index=False, encoding='utf-8-sig')
                
                self.status_var.set(f"成功导出到: {file_path}")
                messagebox.showinfo("成功", f"已成功导出到: {file_path}")
        except Exception as e:
            print(f"导出CSV失败: {str(e)}")
            messagebox.showerror("错误", f"导出失败: {str(e)}")
            self.status_var.set(f"导出失败: {str(e)}")

    def export_to_markdown(self):
        """导出为Markdown文件"""
        try:
            if not self.json_data:
                messagebox.showinfo("提示", "没有数据可导出")
                return
            
            # 防御性编程
            if isinstance(self.json_data, types.GeneratorType):
                self.json_data = list(self.json_data)
            elif hasattr(self.json_data, '__iter__') and not hasattr(self.json_data, '__len__') and not isinstance(self.json_data, (str, bytes, dict)):
                self.json_data = list(self.json_data)
            
            # 生成Markdown内容
            md_content = "# 小说事件汇总\n\n"
            
            for i, data in enumerate(self.json_data, 1):
                title = data.get('title', '未知标题')
                md_content += f"## {i}. {title}\n\n"
                
                # 添加详细版
                md_content += "### 详细版\n\n"
                详细版_list = self._safe_convert_to_list(data.get('详细版', []))
                for j, event in enumerate(详细版_list, 1):
                    try:
                        md_content += f"{event}\n\n"
                    except:
                        md_content += "[数据错误]\n\n"
                
                # 添加简化版
                md_content += "### 简化版\n\n"
                简化版_list = self._safe_convert_to_list(data.get('简化版', []))
                for event in 简化版_list:
                    try:
                        md_content += f"- {event}\n"
                    except:
                        md_content += "- [数据错误]\n"
                md_content += "\n"
                
                # 添加极简版
                md_content += "### 极简版\n\n"
                极简版_list = self._safe_convert_to_list(data.get('极简版', []))
                for event in 极简版_list:
                    try:
                        md_content += f"- {event}\n"
                    except:
                        md_content += "- [数据错误]\n"
                md_content += "\n"
            
            # 保存文件
            default_filename = f"小说事件汇总_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md"
            file_path = filedialog.asksaveasfilename(
                defaultextension=".md",
                filetypes=[("Markdown文件", "*.md")],
                initialfile=default_filename
            )
            
            if file_path:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(md_content)
                self.status_var.set(f"成功导出到: {file_path}")
                messagebox.showinfo("成功", f"已成功导出到: {file_path}")
        except Exception as e:
            print(f"导出Markdown失败: {str(e)}")
            messagebox.showerror("错误", f"导出失败: {str(e)}")
            self.status_var.set(f"导出失败: {str(e)}")

if __name__ == "__main__":
    print("正在启动JSON转换工具...")
    root = tk.Tk()
    print("Tk实例已创建")
    app = JSONConverterApp(root)
    print("应用程序已初始化，启动主循环")
    root.mainloop()
    print("应用程序已关闭")