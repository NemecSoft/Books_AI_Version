import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import json
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, Fill, Protection, PatternFill, GradientFill, Color, colors

class ExcelFormatCopier:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel格式复制工具")
        self.root.geometry("700x500")
        
        # 设置中文字体支持
        try:
            self.style = ttk.Style()
            self.style.configure("TButton", font=('SimHei', 10))
            self.style.configure("TLabel", font=('SimHei', 10))
            print("成功设置中文字体")
        except Exception as e:
            print(f"设置字体时出错: {e}")
        
        # 数据存储
        self.json_files = []
        self.template_path = None
        self.output_path = None
        
        # 创建界面
        self.create_widgets()
    
    def create_widgets(self):
        # 创建主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 模板文件选择区域
        template_frame = ttk.Frame(main_frame, padding="5")
        template_frame.pack(fill=tk.X, side=tk.TOP)
        
        ttk.Label(template_frame, text="Excel模板文件:").pack(side=tk.LEFT, padx=5)
        
        self.template_var = tk.StringVar(value="未选择")
        ttk.Label(template_frame, textvariable=self.template_var, width=50).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(template_frame, text="选择模板", command=self.select_template).pack(side=tk.LEFT, padx=5)
        
        # JSON文件选择区域
        json_frame = ttk.Frame(main_frame, padding="5")
        json_frame.pack(fill=tk.X, side=tk.TOP)
        
        ttk.Button(json_frame, text="添加JSON文件", command=self.add_json_files).pack(side=tk.LEFT, padx=5)
        ttk.Button(json_frame, text="添加文件夹中的JSON", command=self.add_folder_json).pack(side=tk.LEFT, padx=5)
        ttk.Button(json_frame, text="移除选中", command=self.remove_selected).pack(side=tk.LEFT, padx=5)
        ttk.Button(json_frame, text="清空列表", command=self.clear_all).pack(side=tk.LEFT, padx=5)
        
        # JSON文件列表
        ttk.Label(main_frame, text="JSON文件列表:").pack(anchor=tk.W, pady=5)
        
        columns = ("文件名", "路径")
        self.tree = ttk.Treeview(main_frame, columns=columns, show="headings")
        
        # 设置列宽
        self.tree.column("文件名", width=150)
        self.tree.column("路径", width=450)
        
        # 设置列标题
        for col in columns:
            self.tree.heading(col, text=col)
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        
        # 放置表格和滚动条
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # 导出按钮
        export_frame = ttk.Frame(main_frame, padding="5")
        export_frame.pack(fill=tk.X, side=tk.TOP)
        
        ttk.Button(export_frame, text="开始生成Excel", command=self.generate_excel).pack(side=tk.LEFT, padx=5)
        
        # 状态标签
        self.status_var = tk.StringVar(value="就绪")
        status_label = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_label.pack(side=tk.BOTTOM, fill=tk.X)
    
    def select_template(self):
        """选择Excel模板文件"""
        file_path = filedialog.askopenfilename(
            title="选择Excel模板文件",
            filetypes=[("Excel文件", "*.xlsx")]
        )
        
        if file_path:
            self.template_path = file_path
            self.template_var.set(os.path.basename(file_path))
            self.status_var.set(f"已选择模板: {os.path.basename(file_path)}")
    
    def add_json_files(self):
        """添加JSON文件"""
        file_types = [("JSON文件", "*.json"), ("所有文件", "*.*")]
        file_paths = filedialog.askopenfilenames(title="选择JSON文件", filetypes=file_types)
        
        if file_paths:
            self.process_json_files(file_paths)
    
    def add_folder_json(self):
        """添加文件夹中的所有JSON文件"""
        folder_path = filedialog.askdirectory(title="选择包含JSON文件的文件夹")
        
        if folder_path:
            file_paths = []
            for root, _, files in os.walk(folder_path):
                for file in files:
                    if file.lower().endswith('.json'):
                        file_paths.append(os.path.join(root, file))
            
            if file_paths:
                self.process_json_files(file_paths)
            else:
                messagebox.showinfo("提示", f"文件夹 '{folder_path}' 中未找到JSON文件")
    
    def process_json_files(self, file_paths):
        """处理JSON文件并添加到列表中"""
        success_count = 0
        error_count = 0
        
        for file_path in file_paths:
            if file_path in [item[1] for item in self.json_files]:
                continue  # 跳过已添加的文件
            
            try:
                # 验证文件是否为有效的JSON文件
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                # 验证必要字段
                if "基础信息" not in data or "标题" not in data["基础信息"]:
                    raise ValueError("JSON结构不符合要求，缺少必要字段")
                
                if "极简版" not in data or "简化版" not in data or "详细版" not in data:
                    raise ValueError("JSON结构不符合要求，缺少必要的事件列表字段")
                
                # 添加到列表
                file_name = os.path.basename(file_path)
                self.json_files.append((file_name, file_path))
                self.tree.insert("", tk.END, values=(file_name, file_path))
                success_count += 1
                
            except json.JSONDecodeError:
                error_count += 1
            except Exception as e:
                print(f"处理文件 {file_path} 时出错: {str(e)}")
                error_count += 1
        
        self.status_var.set(f"添加完成 - 成功: {success_count}, 失败: {error_count}")
    
    def remove_selected(self):
        """移除选中的JSON文件"""
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showinfo("提示", "请先选择要移除的文件")
            return
        
        removed_count = 0
        for item in selected_items:
            file_path = self.tree.item(item, "values")[1]
            # 查找并移除对应的文件
            for i, (_, path) in enumerate(self.json_files):
                if path == file_path:
                    del self.json_files[i]
                    removed_count += 1
                    break
            # 从树视图中删除
            self.tree.delete(item)
        
        self.status_var.set(f"已移除 {removed_count} 个文件")
    
    def clear_all(self):
        """清空所有JSON文件"""
        if messagebox.askyesno("确认", "确定要清空所有文件吗？"):
            self.tree.delete(*self.tree.get_children())
            self.json_files = []
            self.status_var.set("已清空所有文件")
    
    def generate_excel(self):
        """生成Excel文件，从模板第一行复制格式到所有数据行"""
        try:
            # 检查必要条件
            if not self.template_path:
                messagebox.showinfo("提示", "请先选择Excel模板文件")
                return
            
            if not self.json_files:
                messagebox.showinfo("提示", "请先添加JSON文件")
                return
            
            # 准备数据
            data_to_export = []
            for _, file_path in self.json_files:
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                # 提取数据
                title = data["基础信息"]["标题"]
                
                # 处理各字段
                极简版_list = data.get("极简版", [])
                if not isinstance(极简版_list, list):
                    极简版_list = [极简版_list] if 极简版_list is not None else []
                    
                简化版_list = data.get("简化版", [])
                if not isinstance(简化版_list, list):
                    简化版_list = [简化版_list] if 简化版_list is not None else []
                    
                详细版_list = data.get("详细版", [])
                if not isinstance(详细版_list, list):
                    详细版_list = [详细版_list] if 详细版_list is not None else []
                
                # 生成字符串内容
                极简版_str = "\n".join([f"{i}. {str(event)}" for i, event in enumerate(极简版_list, 1)] if 极简版_list else [])
                简化版_str = "\n".join([f"{i}. {str(event)}" for i, event in enumerate(简化版_list, 1)] if 简化版_list else [])
                详细版_str = "\n".join([f"{i}. {str(event)}" for i, event in enumerate(详细版_list, 1)] if 详细版_list else [])
                
                data_to_export.append([
                    title,
                    极简版_str,
                    简化版_str,
                    详细版_str
                ])
            
            # 创建DataFrame
            df = pd.DataFrame(data_to_export, columns=["标题", "极简版", "简化版", "详细版"])
            
            # 选择输出文件路径
            default_filename = f"小说事件汇总_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx")],
                initialfile=default_filename
            )
            
            if file_path:
                self.status_var.set("正在生成Excel文件...")
                self._export_with_template_format(df, self.template_path, file_path)
                
                self.status_var.set(f"成功生成Excel文件: {file_path}")
                messagebox.showinfo("成功", f"已成功生成Excel文件: {file_path}")
        
        except Exception as e:
            print(f"生成Excel失败: {str(e)}")
            messagebox.showerror("错误", f"生成失败: {str(e)}")
            self.status_var.set(f"生成失败: {str(e)}")
    
    def _export_with_template_format(self, df, template_path, output_path):
        """使用模板格式导出Excel，从第一行复制格式到所有数据行"""
        # 加载模板文件
        template_book = load_workbook(template_path, data_only=True)
        
        # 获取第一个可见工作表
        visible_sheets = [sheet for sheet in template_book.sheetnames if template_book[sheet].sheet_state == 'visible']
        if not visible_sheets:
            raise Exception("模板文件中没有可见的工作表")
        
        template_sheet_name = visible_sheets[0]
        template_worksheet = template_book[template_sheet_name]
        
        # 清除工作表中的所有内容，但保留格式
        max_row = template_worksheet.max_row
        max_col = template_worksheet.max_column
        
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                template_worksheet.cell(row=row_idx, column=col_idx).value = None
        
        # 确保至少有一行作为模板行
        if max_row < 1:
            max_row = 1
        
        if max_col < 4:
            max_col = 4
        
        # 从第一行读取格式
        first_row_styles = []
        for col_idx in range(1, max_col + 1):
            cell = template_worksheet.cell(row=1, column=col_idx)
            # 保存单元格的所有样式属性
            cell_style = {
                'font': cell.font,
                'fill': cell.fill,
                'border': cell.border,
                'alignment': cell.alignment,
                'number_format': cell.number_format,
                'protection': cell.protection
            }
            first_row_styles.append(cell_style)
        
        # 保存第一行的列宽和行高
        column_widths = []
        for col_idx in range(1, max_col + 1):
            column_letter = chr(64 + col_idx)
            if column_letter in template_worksheet.column_dimensions:
                column_widths.append(template_worksheet.column_dimensions[column_letter].width)
            else:
                column_widths.append(None)
        
        first_row_height = template_worksheet.row_dimensions[1].height if 1 in template_worksheet.row_dimensions else None
        
        # 写入数据并应用样式
        for row_idx, row_data in enumerate(df.itertuples(index=False), start=1):
            for col_idx in range(1, min(len(row_data) + 1, max_col + 1)):
                # 获取单元格
                cell = template_worksheet.cell(row=row_idx, column=col_idx)
                
                # 设置值
                cell.value = row_data[col_idx - 1] if (col_idx - 1) < len(row_data) else ""
                
                # 应用第一行的样式
                if (col_idx - 1) < len(first_row_styles):
                    cell_style = first_row_styles[col_idx - 1]
                    cell.font = cell_style['font']
                    cell.fill = cell_style['fill']
                    cell.border = cell_style['border']
                    
                    # 确保文本换行
                    if cell_style['alignment']:
                        new_alignment = Alignment(
                            horizontal=cell_style['alignment'].horizontal if hasattr(cell_style['alignment'], 'horizontal') else 'general',
                            vertical=cell_style['alignment'].vertical if hasattr(cell_style['alignment'], 'vertical') else 'bottom',
                            textRotation=cell_style['alignment'].textRotation if hasattr(cell_style['alignment'], 'textRotation') else 0,
                            wrapText=True,
                            shrinkToFit=cell_style['alignment'].shrinkToFit if hasattr(cell_style['alignment'], 'shrinkToFit') else False,
                            indent=cell_style['alignment'].indent if hasattr(cell_style['alignment'], 'indent') else 0,
                            relativeIndent=cell_style['alignment'].relativeIndent if hasattr(cell_style['alignment'], 'relativeIndent') else 0,
                            justifyLastLine=cell_style['alignment'].justifyLastLine if hasattr(cell_style['alignment'], 'justifyLastLine') else False,
                            readingOrder=cell_style['alignment'].readingOrder if hasattr(cell_style['alignment'], 'readingOrder') else 0
                        )
                    else:
                        new_alignment = Alignment(wrap_text=True)
                    
                    cell.alignment = new_alignment
                    cell.number_format = cell_style['number_format']
                    cell.protection = cell_style['protection']
            
            # 设置行高
            if first_row_height:
                template_worksheet.row_dimensions[row_idx].height = first_row_height
        
        # 确保所有列宽设置正确
        for col_idx in range(1, max_col + 1):
            if (col_idx - 1) < len(column_widths) and column_widths[col_idx - 1]:
                column_letter = chr(64 + col_idx)
                template_worksheet.column_dimensions[column_letter].width = column_widths[col_idx - 1]
        
        # 保存文件
        template_book.save(output_path)
        print(f"成功导出Excel文件: {output_path}")

if __name__ == "__main__":
    print("正在启动Excel格式复制工具...")
    root = tk.Tk()
    print("Tk实例已创建")
    app = ExcelFormatCopier(root)
    print("应用程序已初始化，启动主循环")
    root.mainloop()
    print("应用程序已关闭")
